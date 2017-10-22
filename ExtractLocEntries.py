#!/usr/bin/env python3
#! -*- coding:utf-8 -*-

import os
import pathlib
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def addWorkSheet(ws, resPath):

    ws['A2']='Source'
    ws['B2']='KeyName'
    ws['C2']='Notes'
    ws['D2']='English'
    ws['E2']='Chinese'
    ws['F2']='Hindi'
    ws['G2']='Indonesia'

    soup = BeautifulSoup(open(resPath, encoding='utf-8'), 'html.parser')
    resources = soup.resources.findAll('string')
    ln = 3
    for resource in resources:
        posSource = 'A%d' % ln
        ws[posSource] = str(resource)
        posKeyName= 'B%d' % ln
        ws[posKeyName] = resource['name']
        posEnglish = 'D%d' % ln
        ws[posEnglish] = resource.string
        ln += 1
    ws['A1'] = 'count'
    ws['B1'] = len(resources)

    ws['C1'] = 'Path to strings.xml:'
    ws['D1'] = resPath # Todo: relative path instead of real path

def extractLocEntriesToExcel():
    prjRootPath = '/Users/weiqi/alisrc/Global/MobileSecurity'  # pass in the project dir

    resFiles = []
    for path,subdir,files in os.walk(prjRootPath):
        for name in files:
            if name.lower() == 'strings.xml' and path.endswith('values'):
                #resFiles.append(pathlib.PurePath(path, name))
                resFiles.append(os.path.join(path, name))

    wb = Workbook()
    wsSummary = wb.active
    wsSummary.title = "Summary"

    cnt = 0
    for item in resFiles:
        wsCurrent = wb.create_sheet()
        addWorkSheet(wsCurrent, item)
        cnt = cnt + 1

    wsSummary['A10']= 'Number of Sheet:'
    wsSummary['B10']= cnt
    wsSummary['A9'] = 'Project Path'
    wsSummary['B9'] = prjRootPath
    wb.save("test.xlsx")

def RestoreExcel():
    prjRootPath = '/Users/weiqi/alisrc/Global/MobileSecurity'
    pathToExcel = 'test.xlsx'

    wb = load_workbook(pathToExcel)
    wsSummary = wb.get_sheet_by_name('Summary')
    origPrjPath = wsSummary['B9'].value
    cntSheets = int(wsSummary['B10'].value)
    print(origPrjPath)
    print(cntSheets)

    for ws in wb.worksheets:
        if ws.title.lower() == 'summary':
            continue
        cnt = int(ws['B1'].value)

    #
    # wb = load_workbook('sample.xlsx')
    # ws = wb.active
    # cnt = int(ws['B1'].value)
    # print(cnt)
    # soup = BeautifulSoup(open('strings_call_recorder.xml', encoding='utf-8'), 'html.parser')
    # resources = soup.resources.findAll('string')
    #
    # # Load translations
    # translations = {}
    # for i in range(3, cnt + 3):
    #     item1 = ws.cell(row=i, column=2)
    #     item2 = ws.cell(row=i, column=5)
    #     translations[item1.value] = item2.value
    #
    # for resource in resources:
    #     key = resource['name']
    #     if key == 'language':  # no need to translate 'language'
    #         continue
    #     resource.string = str(translations[key])
    # with open('strings_translated.xml', 'w', encoding='utf-8') as f:
    #     f.write(str(soup))


if __name__ == '__main__':
    # extractLocEntries()
    RestoreExcel()