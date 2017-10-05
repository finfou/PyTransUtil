#!/usr/bin/env python3
# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup

if __name__ == '__main__':
    wb = load_workbook('sample.xlsx')
    ws = wb.active
    cnt = int(ws['B1'].value)
    print(cnt)
    soup = BeautifulSoup(open('strings_call_recorder.xml', encoding='utf-8'), 'html.parser')
    resources = soup.resources.findAll('string')

    # Load translations
    translations = {}
    for i in range(3,cnt+3):
        item1 = ws.cell(row=i, column=2)
        item2 = ws.cell(row=i, column=5)
        translations[item1.value]=item2.value

    for resource in resources:
        key = resource['name']
        if key == 'language': # no need to translate 'language'
            continue
        resource.string = str(translations[key])
    with open('strings_translated.xml', 'w', encoding='utf-8') as f:
        f.write(str(soup))


