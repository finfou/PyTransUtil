#!/usr/bin/env python3
# -*- coding:utf-8 -*-
from openpyxl import Workbook
from bs4 import BeautifulSoup


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active

    # WriteHead(wb)
    ws['A2']='Source'
    ws['B2']='KeyName'
    ws['C2']='Notes'
    ws['D2']='English'
    ws['E2']='Chinese'
    ws['F2']='Hindi'
    ws['G2']='Indonesia'

    soup = BeautifulSoup(open('strings_call_recorder.xml', encoding='utf-8'))
    resources = soup.resources.findAll('string')
    ln = 3
    for resource in resources:
        posSource = 'A%d' % ln
        ws[posSource] = str(resource)
        posKeyName= 'B%d' % ln
        ws[posKeyName] = resource['name']
        posEnglish = 'D%d' % ln
        ws[posEnglish] = resource.string
        ln += 1
    ws['A1'] = 'count'
    ws['B1'] = len(resources)

    wb.save("sample.xlsx")
